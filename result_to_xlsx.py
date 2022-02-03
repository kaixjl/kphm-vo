import os
import os.path
from collections import OrderedDict
from itertools import chain, product
from typing import Dict, List, Tuple
import pickle
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

def parse_eval_pose_result(filepath):
    # type: (str) -> List[Dict]
    eval_pose_result = []
    keys = ["seq_id",
            "translation-length",
            "rotation-length",
            "ATE",
            "RPE",
            "RPE(deg)"]
    with open(filepath, 'r') as f:
        rst = f.read().splitlines()
        i = 0
        while i < len(rst) and rst[i].startswith("Sequence"):
            metrics = [l.split(":") for l in rst[i:i+6]]
            metric_dict = {}
            for j in range(len(keys)):
                metric_dict[keys[j]] = metrics[j][1].strip()
            try:
                metric_dict["seq_id"] = "{:02d}".format(int(metric_dict["seq_id"]))
            except:
                pass
            eval_pose_result.append(metric_dict)
            i += 7

    return eval_pose_result

def parse_hyperparams_string(s):
    # type: (str) -> dict
    s = s.splitlines()
    s = [l.strip().split(":") for l in s if len(l.strip()) > 0]
    s = {k:v.strip() for k, v in s}
    return s

def parse_hyperparameters(tag, config):
    if os.path.exists(os.path.join("archive", tag, "result", config, "hyperparams.txt")):
        with open(os.path.join("archive", tag, "result", config, "hyperparams.txt"), 'r') as f:
            s = f.read()
        hparams = parse_hyperparams_string(s)
    elif os.path.exists(os.path.join("archive", tag, "records", config, "hyperparams.pickle")):
        with open(os.path.join("archive", tag, "records", config, "hyperparams.pickle"), 'rb') as f:
            s = pickle.load(f) # type: dict
        hparams = {k: str(v) for k, v in s.items()}
    else:
        hparams = {}

    return hparams

def write_result(output_filepath):
    rsts = {}
    hparamss = {}
    tags = os.listdir("archive")
    for tag in tags:
        path_result = os.path.join("archive", tag, "result")
        if not os.path.exists(path_result):
            continue
        configs = os.listdir(path_result)
        configs.sort()
        for config in configs:
            path_eval_pose_result = os.path.join(path_result, config, "result.txt")

            if os.path.exists(path_eval_pose_result):
                kitti_metrics = parse_eval_pose_result(path_eval_pose_result)
            else:
                kitti_metrics = {}

            hparams = parse_hyperparameters(tag, config)

            rsts["{}/{}".format(tag, config)] = kitti_metrics
            hparamss["{}/{}".format(tag, config)] = hparams

    metric_seqs = set(metric["seq_id"] for metric in chain.from_iterable(v for v in rsts.values()))
    metric_titles = set(chain.from_iterable(metric.keys() for metric in chain.from_iterable(v for v in rsts.values())))
    if "seq_id" in metric_titles:
        metric_titles.remove("seq_id")

    titles = ["archive/config"]
    titles_kitti_metrics = ["{}-{}".format(seq_id, title) for seq_id, title in product(metric_seqs, metric_titles)]
    titles_kitti_metrics.sort()
    titles.extend(titles_kitti_metrics)
    titles_hparams_t = set()
    for hparams in hparamss.values():
        titles_hparams_t.update(hparams.keys())
    titles_hparams_t = list(titles_hparams_t)
    titles_hparams_t.sort()
    titles_hparams = ["NETWORK"]
    titles_hparams.extend([t for t in titles_hparams_t if t.startswith("WITH_")])
    titles_hparams.extend(["IMG_HEIGHT", "IMG_WIDTH", "SPLIT_TAG", "BATCH_SIZE", "EPOCH_SIZE"])
    titles_hparams_set = set(titles_hparams)
    titles_hparams_t = [t for t in titles_hparams_t if t not in titles_hparams_set]
    titles_hparams.extend(titles_hparams_t)
    titles.extend(titles_hparams)

    rows = []
    for k, kitti_metrics in rsts.items():
        row = {"archive/config": k}
        for kitti_metric in kitti_metrics:
            seq_id = kitti_metric["seq_id"]
            for title, kitti_metric_v in kitti_metric.items():
                if title == "seq_id":
                    continue
                row["{}-{}".format(seq_id, title)] = float(kitti_metric_v)
        for key, value in hparamss[k].items():
            row[key] = value
        rows.append(row)

    wb = Workbook()
    ws = wb.active # type: Worksheet
    
    # with open(output_filepath, 'w') as f:
    ws.append(titles)
    # f.write("{}\n".format(",".join(titles)))
    for row in rows:
        row_txt = []
        for title in titles:
            if title in row.keys():
                row_txt.append(row[title])
            else:
                row_txt.append("")
        ws.append(row_txt)
        # row_txt = ",".join(row_txt)
        # f.write("{}\n".format(row_txt))

    def itoABC(i):
        n = 1
        p = 26
        i -= 1
        while i >= p:
            n += 1
            p += 26**n
        for j in range(n - 1):
            i -= 26**(j+1)
        s = ""
        while True:
            t = i % 26
            i = i // 26
            s = chr(t + 65) + s
            if i == 0:
                break
        s = "A"*(n-len(s)) + s
        return s

        
    for i in range(1,len(titles)+1):
        range_str = '{}2:{}{}'.format(itoABC(i), itoABC(i), len(rows)+1)
        ws.conditional_formatting.add(range_str,
            ColorScaleRule(start_type='min', start_color='5A8AC6',
            mid_type='percentile', mid_value=50, mid_color='FFFFFF',
            end_type='max', end_color='F8696B')
            )
    
    for i, j in product(range(1, len(rows)+1), range(1, len(titles))):
        ws.cell(row=i, column=j).number_format = '0.0000' 
                    
    ws.column_dimensions['A'].width = 60
    wb.save(output_filepath)


def main():

    write_result("output_result.xlsx")

    pass
# } def main

if __name__=="__main__":
    main()